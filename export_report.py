import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image

class ReportExporter:
    def __init__(self):
        self.wb = Workbook()
        
    def export_ahp_results(self, criteria_names, comparison_matrix, weights, cr, ci):
        """Export AHP results to Excel"""
        ws = self.wb.create_sheet("AHP Analysis")
        
        # Title
        ws['A1'] = "AHP Analysis Results"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Criteria Names
        ws['A3'] = "Criteria Names"
        ws['A3'].font = Font(bold=True)
        for i, name in enumerate(criteria_names, 1):
            ws[f'A{i+3}'] = name
        
        # Pairwise Comparison Matrix
        ws['C3'] = "Pairwise Comparison Matrix"
        ws['C3'].font = Font(bold=True)
        
        num_criteria = len(criteria_names)
        matrix_title_row = 3
        matrix_header_row = matrix_title_row 
        matrix_first_data_row = matrix_header_row + 1
        matrix_last_data_row = matrix_header_row + num_criteria

        # Matrix headers
        for i, name in enumerate(criteria_names, 1):
            ws[f'{get_column_letter(i+3)}3'] = name # Column D onwards for criteria names in header
            ws[f'C{i+3}'] = name # Column C for criteria names as row labels
        
        # Matrix values
        for i in range(num_criteria):
            for j in range(num_criteria):
                ws[f'{get_column_letter(j+4)}{i+4}'] = comparison_matrix[i][j]
        
        section_buffer = 3
        weights_consistency_start_row = matrix_last_data_row + section_buffer
        
        # Weights
        weights_title_row = weights_consistency_start_row
        weights_header_row = weights_title_row + 1
        weights_first_data_row = weights_header_row + 1

        ws[f'A{weights_title_row}'] = "Criteria Weights"
        ws[f'A{weights_title_row}'].font = Font(bold=True)
        ws[f'A{weights_header_row}'] = "Criteria"
        ws[f'B{weights_header_row}'] = "Weight"
        ws[f'A{weights_header_row}'].font = Font(bold=True)
        ws[f'B{weights_header_row}'].font = Font(bold=True)
        
        for idx, (name, weight) in enumerate(zip(criteria_names, weights), 1):
            current_data_row = weights_header_row + idx
            ws[f'A{current_data_row}'] = name
            ws[f'B{current_data_row}'] = weight
        
        # Consistency
        consistency_title_row = weights_consistency_start_row # Same start row as weights, but different columns
        ws[f'D{consistency_title_row}'] = "Consistency Analysis"
        ws[f'D{consistency_title_row}'].font = Font(bold=True)
        ws[f'D{consistency_title_row + 1}'] = "Consistency Index (CI)"
        ws[f'E{consistency_title_row + 1}'] = ci
        ws[f'D{consistency_title_row + 2}'] = "Consistency Ratio (CR)"
        ws[f'E{consistency_title_row + 2}'] = cr
        
        # Create bar chart for weights
        chart = BarChart()
        chart.title = "Criteria Weights"
        chart.y_axis.title = "Weight"
        chart.x_axis.title = "Criteria"
        
        data = Reference(ws, min_col=2, min_row=weights_header_row, max_row=weights_header_row + num_criteria)
        cats = Reference(ws, min_col=1, min_row=weights_first_data_row, max_row=weights_first_data_row + num_criteria -1) # Equivalent to weights_header_row + num_criteria
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        matrix_last_content_col_idx = 3 + num_criteria # Matrix content goes from col C + num_criteria columns
        chart_start_col_letter = get_column_letter(matrix_last_content_col_idx + 2) # Buffer of 2 columns
        chart_anchor_cell = f'{chart_start_col_letter}{matrix_title_row}' # Align with matrix title row

        ws.add_chart(chart, chart_anchor_cell)
        
        # Adjust column widths (first 9 columns, A-I)
        for col_idx_1_based in range(1, 10):
            ws.column_dimensions[get_column_letter(col_idx_1_based)].width = 15
    
    def export_topsis_results(self, alternatives, criteria_names, performance_matrix, 
                            benefit_criteria, weights, scores, rankings):
        """Export TOPSIS results to Excel"""
        ws = self.wb.create_sheet("TOPSIS Analysis")
        
        # Title
        ws['A1'] = "TOPSIS Analysis Results"
        ws['A1'].font = Font(size=14, bold=True)

        num_criteria = len(criteria_names)
        num_alternatives = len(alternatives)
        section_buffer = 3

        # Performance Matrix
        perf_matrix_title_row = 3
        ws[f'A{perf_matrix_title_row}'] = "Performance Matrix"
        ws[f'A{perf_matrix_title_row}'].font = Font(bold=True)
        
        perf_matrix_header_row = perf_matrix_title_row + 1
        perf_matrix_first_data_row = perf_matrix_header_row + 1
        perf_matrix_last_data_row = perf_matrix_header_row + num_alternatives

        # Headers for Performance Matrix
        ws[f'A{perf_matrix_header_row}'] = "Alternative"
        ws[f'A{perf_matrix_header_row}'].font = Font(bold=True)
        for i, name in enumerate(criteria_names, 1):
            col_letter = get_column_letter(i + 1) # Criteria start from column B
            ws[f'{col_letter}{perf_matrix_header_row}'] = name
            ws[f'{col_letter}{perf_matrix_header_row}'].font = Font(bold=True)
        
        # Matrix values
        for i, alt in enumerate(alternatives):
            current_row = perf_matrix_first_data_row + i
            ws[f'A{current_row}'] = alt
            for j in range(num_criteria):
                ws[f'{get_column_letter(j+2)}{current_row}'] = performance_matrix[i][j]
        
        # Criteria Types and Weights Section
        types_weights_start_row = perf_matrix_last_data_row + section_buffer
        types_weights_title_row = types_weights_start_row
        types_weights_header_row = types_weights_title_row + 1
        types_weights_first_data_row = types_weights_header_row + 1

        # Criteria Types
        ws[f'A{types_weights_title_row}'] = "Criteria Types"
        ws[f'A{types_weights_title_row}'].font = Font(bold=True)
        ws[f'A{types_weights_header_row}'] = "Criteria"
        ws[f'B{types_weights_header_row}'] = "Type"
        ws[f'A{types_weights_header_row}'].font = Font(bold=True)
        ws[f'B{types_weights_header_row}'].font = Font(bold=True)
        
        for idx, (name, is_benefit) in enumerate(zip(criteria_names, benefit_criteria), 1):
            current_data_row = types_weights_header_row + idx
            ws[f'A{current_data_row}'] = name
            ws[f'B{current_data_row}'] = "Benefit" if is_benefit else "Cost"
        
        # Criteria Weights
        ws[f'D{types_weights_title_row}'] = "Criteria Weights" # Same title row as Types, different columns
        ws[f'D{types_weights_title_row}'].font = Font(bold=True)
        ws[f'D{types_weights_header_row}'] = "Criteria"
        ws[f'E{types_weights_header_row}'] = "Weight"
        ws[f'D{types_weights_header_row}'].font = Font(bold=True)
        ws[f'E{types_weights_header_row}'].font = Font(bold=True)
        
        for idx, (name, weight) in enumerate(zip(criteria_names, weights), 1):
            current_data_row = types_weights_header_row + idx
            ws[f'D{current_data_row}'] = name
            ws[f'E{current_data_row}'] = weight
        
        last_row_of_types_weights_section = types_weights_header_row + num_criteria

        # Final Rankings Section
        rankings_start_row = last_row_of_types_weights_section + section_buffer
        rankings_title_row = rankings_start_row
        rankings_header_row = rankings_title_row + 1
        rankings_first_data_row = rankings_header_row + 1

        ws[f'A{rankings_title_row}'] = "Final Rankings"
        ws[f'A{rankings_title_row}'].font = Font(bold=True)
        ws[f'A{rankings_header_row}'] = "Alternative"
        ws[f'B{rankings_header_row}'] = "Score"
        ws[f'C{rankings_header_row}'] = "Rank"
        ws[f'A{rankings_header_row}'].font = Font(bold=True)
        ws[f'B{rankings_header_row}'].font = Font(bold=True)
        ws[f'C{rankings_header_row}'].font = Font(bold=True)
        
        # Sort alternatives by rank
        ranked_data = sorted(zip(alternatives, scores, rankings), key=lambda x: x[2])
        for idx, (alt, score, rank_val) in enumerate(ranked_data, 1): # Renamed rank to rank_val
            current_data_row = rankings_header_row + idx
            ws[f'A{current_data_row}'] = alt
            ws[f'B{current_data_row}'] = score
            ws[f'C{current_data_row}'] = rank_val
        
        # Create bar chart for rankings
        chart = BarChart()
        chart.title = "Alternative Rankings"
        chart.y_axis.title = "Score"
        chart.x_axis.title = "Alternative"
        
        data = Reference(ws, min_col=2, min_row=rankings_header_row, max_row=rankings_header_row + num_alternatives)
        cats = Reference(ws, min_col=1, min_row=rankings_first_data_row, max_row=rankings_first_data_row + num_alternatives -1) # Equivalent to rankings_header_row + num_alternatives
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        perf_matrix_last_col_idx = 1 + num_criteria # Performance matrix has 1 (Alternative) + num_criteria columns
        chart_start_col_letter = get_column_letter(perf_matrix_last_col_idx + 2) # Buffer of 2 columns
        chart_anchor_cell = f'{chart_start_col_letter}{perf_matrix_title_row}' # Align with TOPSIS title row

        ws.add_chart(chart, chart_anchor_cell)
        
        # Adjust column widths
        for col_idx_1_based in range(1, 10): # Adjust first 9 columns (A-I)
            ws.column_dimensions[get_column_letter(col_idx_1_based)].width = 15
    
    def save_report(self, filename):
        """Save the Excel report"""
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        
        self.wb.save(filename) 